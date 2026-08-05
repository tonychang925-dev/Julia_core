"""Julia File System Capability — eyes and hands for the digital world.

Four tools: list, search, read (all formats), write.
Principle: return raw content, never summarize. LLM does the understanding.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


class FileSystemTool:
    """Unified filesystem capability. LLM decides actions. Runtime provides data."""

    @staticmethod
    def list_directory(path: str) -> str:
        """List directory contents with metadata."""
        p = Path(path).expanduser()
        if not p.exists():
            return f"目录不存在: {path}"
        items = []
        for item in sorted(p.iterdir()):
            if item.name.startswith('.'):
                continue
            stat = item.stat()
            mtime = datetime.fromtimestamp(stat.st_mtime).strftime("%m/%d %H:%M")
            size = stat.st_size
            if item.is_dir():
                items.append(f"  📁 {item.name}/ ({mtime})")
            else:
                size_str = f"{size:,}B" if size < 1024 else f"{size/1024:.0f}KB" if size < 1024*1024 else f"{size/1024/1024:.1f}MB"
                items.append(f"  📄 {item.name} ({size_str}, {mtime})")
        header = f"{path}/ ({len(items)} items)"
        return header + "\n" + "\n".join(items[:50])

    @staticmethod
    def search_files(pattern: str, directory: str = "/Users/admin", max_depth: int = 3) -> str:
        """Search files by name or content pattern."""
        results = []
        search_dir = Path(directory).expanduser()
        if not search_dir.exists():
            return f"目录不存在: {directory}"
        # Shallow search first, then deeper
        for depth in range(1, max_depth + 1):
            glob_pattern = "/".join(["*"] * depth)
            for path in search_dir.glob(f"{glob_pattern}*{pattern}*"):
                if path.name.startswith('.') or '__pycache__' in str(path) or 'node_modules' in str(path):
                    continue
                results.append(str(path))
                if len(results) >= 20:
                    break
            if len(results) >= 20:
                break
        if not results:
            return f"未找到匹配 '{pattern}' 的文件"
        return "\n".join(results[:20])

    @staticmethod
    def read_file(path: str, max_chars: int = 8000) -> str:
        """Read any supported file format. Returns raw content."""
        p = Path(path).expanduser()
        if not p.exists():
            return f"文件不存在: {path}"

        suffix = p.suffix.lower()

        # Plain text formats
        if suffix in ('.md', '.txt', '.py', '.js', '.ts', '.json', '.yaml', '.yml',
                       '.html', '.css', '.sh', '.toml', '.cfg', '.ini', '.xml', '.csv'):
            try:
                content = p.read_text(encoding='utf-8', errors='ignore')
                return f"=== {p.name} ===\n{content[:max_chars]}"
            except Exception:
                return f"无法读取: {p.name}"

        # PDF
        if suffix == '.pdf':
            try:
                import subprocess
                result = subprocess.run(
                    ["pdftotext", "-layout", str(p), "-"],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode == 0 and result.stdout.strip():
                    return f"=== {p.name} (PDF) ===\n{result.stdout[:max_chars]}"
            except Exception:
                pass
            return f"=== {p.name} (PDF) ===\n[需要安装 pdftotext 来读取 PDF 文件]"

        # DOCX
        if suffix == '.docx':
            try:
                from docx import Document
                doc = Document(str(p))
                text = "\n".join(p.text for p in doc.paragraphs)
                return f"=== {p.name} (DOCX) ===\n{text[:max_chars]}"
            except ImportError:
                return f"=== {p.name} (DOCX) ===\n[需要安装 python-docx: pip install python-docx]"
            except Exception:
                return f"无法读取DOCX: {p.name}"

        # XLSX
        if suffix == '.xlsx':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                lines = []
                for sheet_name in wb.sheetnames[:3]:
                    ws = wb[sheet_name]
                    lines.append(f"\n--- Sheet: {sheet_name} ---")
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                        row_count += 1
                        if row_count >= 50:
                            lines.append("... (truncated)")
                            break
                wb.close()
                return "\n".join(lines)[:max_chars]
            except ImportError:
                return f"=== {p.name} (XLSX) ===\n[需要安装 openpyxl: pip install openpyxl]"
            except Exception as e:
                return f"无法读取XLSX: {p.name} — {e}"

        # Image
        if suffix in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'):
            size = p.stat().st_size
            return f"=== {p.name} (IMAGE) ===\n图片文件，{size/1024:.0f}KB。使用 vision_analyze 工具来查看图片内容。"

        # Unknown format
        try:
            content = p.read_text(encoding='utf-8', errors='ignore')
            return f"=== {p.name} ===\n{content[:max_chars]}"
        except Exception:
            return f"无法读取文件: {p.name} (不支持的格式: {suffix})"

    @staticmethod
    def write_file(path: str, content: str) -> str:
        """Write content to a file. Creates parent directories if needed."""
        p = Path(path).expanduser()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding='utf-8')
        return f"已写入: {p} ({len(content)} 字符)"


# ── Tool Registry Integration ───────────────────────────────────────────────

def register_filesystem_tools(registry):
    """Register all filesystem tools in the tool registry."""
    from julia_core.capability.tool_protocol import ToolSchema, ToolCategory

    registry.register(
        ToolSchema(
            name="list_directory",
            description="列出目录内容。当你想了解项目结构、查看文件或整理桌面时使用。",
            category=ToolCategory.FILE,
            parameters={"path": "目录路径"},
            example="list_directory(path='/Users/admin/Desktop')",
        ),
        lambda path="/Users/admin": FileSystemTool.list_directory(path),
    )

    registry.register(
        ToolSchema(
            name="search_files",
            description="按文件名搜索文件。支持模糊匹配。当Tony提到某个文件但你不知道位置时使用。",
            category=ToolCategory.FILE,
            parameters={"pattern": "文件名关键词", "directory": "搜索目录（可选）"},
            example="search_files(pattern='julia', directory='/Users/admin')",
        ),
        lambda pattern="", directory="/Users/admin": FileSystemTool.search_files(pattern, directory),
    )

    registry.register(
        ToolSchema(
            name="read_file",
            description="读取文件内容。支持 PDF, DOCX, XLSX, MD, TXT, PY, JSON 等格式。当Tony让你看文件时使用。",
            category=ToolCategory.FILE,
            parameters={"path": "文件路径"},
            example="read_file(path='/Users/admin/julia_core/README.md')",
        ),
        lambda path="": FileSystemTool.read_file(path),
    )

    registry.register(
        ToolSchema(
            name="write_file",
            description="创建或覆盖文件。当Tony让你保存内容、写日记、创建文档时使用。需要LLM确认后再调用。",
            category=ToolCategory.FILE,
            parameters={"path": "文件路径", "content": "文件内容"},
            example="write_file(path='/Users/admin/notes/today.md', content='# 今日笔记')",
        ),
        lambda path="", content="": FileSystemTool.write_file(path, content),
    )
